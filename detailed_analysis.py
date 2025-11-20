import pandas as pd
import json
from openpyxl import load_workbook

# 读取Excel文件详细结构
excel_file = r'E:\wangchen\mytest\dingding-demo\精益数据统计.xlsx'

print("Excel文件详细分析:")
print("=" * 60)

try:
    # 使用openpyxl读取
    wb = load_workbook(excel_file)

    for sheet_name in wb.sheetnames:
        print(f"\n\n工作表: {sheet_name}")
        print("-" * 40)
        sheet = wb[sheet_name]

        # 读取所有数据
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        # 打印前20行数据
        print(f"总行数: {len(data)}, 总列数: {len(data[0]) if data else 0}")

        for i, row in enumerate(data[:20]):  # 只显示前20行
            if any(cell is not None for cell in row):  # 只显示非空行
                print(f"第{i+1}行: {row}")

        print("...")

except Exception as e:
    print(f"读取Excel文件出错: {e}")
    # 备用方案：使用pandas
    try:
        all_sheets = pd.read_excel(excel_file, sheet_name=None)
        for sheet_name, df in all_sheets.items():
            print(f"\n工作表: {sheet_name}")
            print(df.to_string())
    except Exception as e2:
        print(f"pandas读取也失败: {e2}")

# 详细分析JSON结构
print("\n\n\n钉钉表单JSON详细结构分析:")
print("=" * 60)

with open(r'E:\wangchen\mytest\dingding-demo\processInstanceResult.json', 'r', encoding='utf-8') as f:
    json_data = json.load(f)

# 1. 分析主数据结构
print("1. 主流程实例数据:")
main_fields = ['businessId', 'title', 'createTime', 'status', 'result',
               'originatorUserId', 'originatorDeptId', 'originatorDeptName']
for field in main_fields:
    print(f"   {field}: {json_data.get(field, 'N/A')}")

# 2. 分析表单组件
print("\n2. 表单组件详细分析:")
form_components = json_data.get('formComponentValues', [])
for i, component in enumerate(form_components):
    print(f"   组件{i+1}:")
    print(f"     名称: {component.get('name')}")
    print(f"     类型: {component.get('componentType')}")
    print(f"     ID: {component.get('id')}")
    print(f"     值: {component.get('value')}")
    if component.get('extValue'):
        print(f"     扩展值: {component.get('extValue')}")
    print()

# 3. 分析操作记录
print("3. 操作记录分析:")
operation_records = json_data.get('operationRecords', [])
for i, record in enumerate(operation_records):
    print(f"   记录{i+1}:")
    print(f"     操作: {record.get('showName')}")
    print(f"     类型: {record.get('type')}")
    print(f"     时间: {record.get('date')}")
    print(f"     结果: {record.get('result')}")
    print(f"     用户ID: {record.get('userId')}")
    print(f"     备注: {record.get('remark')}")
    print()

# 4. 分析任务
print("4. 任务分析:")
tasks = json_data.get('tasks', [])
for i, task in enumerate(tasks):
    print(f"   任务{i+1}:")
    print(f"     任务ID: {task.get('taskId')}")
    print(f"     活动ID: {task.get('activityId')}")
    print(f"     用户ID: {task.get('userId')}")
    print(f"     状态: {task.get('status')}")
    print(f"     结果: {task.get('result')}")
    print(f"     创建时间: {task.get('createTime')}")
    print(f"     完成时间: {task.get('finishTime')}")
    print()

# 统计分析
print("5. 数据统计:")
print(f"   表单组件总数: {len(form_components)}")
print(f"   操作记录总数: {len(operation_records)}")
print(f"   任务总数: {len(tasks)}")

# 组件类型统计
component_types = {}
for component in form_components:
    comp_type = component.get('componentType')
    if comp_type not in component_types:
        component_types[comp_type] = []
    component_types[comp_type].append({
        'name': component.get('name'),
        'value': component.get('value')
    })

print("\n   组件类型分布:")
for comp_type, components in component_types.items():
    print(f"     {comp_type}: {len(components)}个")
    for comp in components:
        print(f"       - {comp['name']}: {comp['value'][:50]}{'...' if len(str(comp['value'])) > 50 else ''}")
    print()